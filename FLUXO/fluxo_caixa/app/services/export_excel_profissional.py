"""Exportador Excel profissional usando openpyxl (pandas como apoio opcional).

Cria múltiplas abas, aplica formatação e adiciona gráficos básicos.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from pathlib import Path
from datetime import datetime, date


def gerar_planilha_profissional(lancamentos: list, resumo_mensal: pd.DataFrame, resumo_anual: pd.DataFrame, caminho_saida: Path):
    """Gera arquivo Excel com múltiplas abas, formatação e gráficos."""
    wb = Workbook()

    # Estilos
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    zebra_fill = PatternFill(start_color='F7F9FB', end_color='F7F9FB', fill_type='solid')
    thin = Side(border_style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color='FFFFFF')
    bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    currency_format = 'R$ #,##0.00'
    date_format = 'yyyy-mm-dd'

    # === Dados_Brutos ===
    ws_raw = wb.active
    ws_raw.title = 'Dados_Brutos'

    df_raw = pd.DataFrame(lancamentos)
    col_order = [
        'data', 'tipo', 'categoria', 'subcategoria', 'descricao',
        'valor', 'banco', 'nota_fiscal', 'empresa'
    ]
    if df_raw.empty:
        df_raw = pd.DataFrame(columns=col_order)
    else:
        for col in col_order:
            if col not in df_raw.columns:
                df_raw[col] = ""
        df_raw = df_raw[col_order]

    # Cabeçalho
    for c_idx, col in enumerate(df_raw.columns, start=1):
        cell = ws_raw.cell(row=1, column=c_idx, value=col.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    # Dados
    for r_idx, row in enumerate(df_raw.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_raw.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border

            coluna = df_raw.columns[c_idx - 1]
            if coluna == 'data':
                if isinstance(value, str):
                    try:
                        value_dt = datetime.strptime(value, "%Y-%m-%d")
                        cell.value = value_dt
                        cell.number_format = date_format
                    except ValueError:
                        pass
                elif isinstance(value, (datetime, date)):
                    cell.number_format = date_format
                cell.alignment = align_center
            elif coluna == 'valor':
                try:
                    cell.value = float(value)
                except Exception:
                    pass
                cell.number_format = currency_format
                cell.alignment = align_right
            else:
                cell.alignment = align_left

        if r_idx % 2 == 0:
            for c_idx in range(1, len(df_raw.columns) + 1):
                ws_raw.cell(row=r_idx, column=c_idx).fill = zebra_fill

    # Larguras
    widths = [12, 12, 18, 18, 40, 14, 12, 16, 24]
    for idx, width in enumerate(widths, start=1):
        ws_raw.column_dimensions[get_column_letter(idx)].width = width

    ws_raw.freeze_panes = "A2"
    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(df_raw.columns))}{ws_raw.max_row}"
    try:
        tabela = Table(displayName="TabelaLancamentos", ref=ws_raw.auto_filter.ref)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws_raw.add_table(tabela)
    except Exception:
        pass

    # === Resumo_Mensal ===
    ws_month = wb.create_sheet('Resumo_Mensal')
    if resumo_mensal.empty:
        resumo_mensal = pd.DataFrame(columns=['ano_mes', 'entradas', 'saidas', 'saldo'])
    for r_idx, row in enumerate(dataframe_to_rows(resumo_mensal, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_month.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
            else:
                if c_idx >= 2:
                    cell.number_format = currency_format
                    cell.alignment = align_right
                else:
                    cell.alignment = align_center
    for idx, width in enumerate([12, 16, 16, 16], start=1):
        ws_month.column_dimensions[get_column_letter(idx)].width = width
    ws_month.freeze_panes = "A2"

    # === Resumo_Anual ===
    ws_year = wb.create_sheet('Resumo_Anual')
    if resumo_anual.empty:
        resumo_anual = pd.DataFrame(columns=['ano', 'entradas', 'saidas', 'saldo'])
    for r_idx, row in enumerate(dataframe_to_rows(resumo_anual, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_year.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
            else:
                if c_idx >= 2:
                    cell.number_format = currency_format
                    cell.alignment = align_right
                else:
                    cell.alignment = align_center
    for idx, width in enumerate([10, 16, 16, 16], start=1):
        ws_year.column_dimensions[get_column_letter(idx)].width = width
    ws_year.freeze_panes = "A2"

    # === Indicadores ===
    ws_ind = wb.create_sheet('Indicadores')
    ws_ind.column_dimensions['A'].width = 26
    ws_ind.column_dimensions['B'].width = 18

    ws_ind['A1'] = 'RESUMO EXECUTIVO'
    ws_ind['A1'].font = header_font
    ws_ind['A1'].fill = header_fill
    ws_ind['A1'].alignment = align_center
    ws_ind.merge_cells('A1:B1')

    total_entradas = float(resumo_anual['entradas'].sum()) if 'entradas' in resumo_anual.columns else 0
    total_saidas = float(resumo_anual['saidas'].sum()) if 'saidas' in resumo_anual.columns else 0
    saldo = total_entradas - total_saidas

    indicadores = [
        ("Total Entradas", total_entradas, "70AD47"),
        ("Total Saídas", total_saidas, "E74C3C"),
        ("Saldo Total", saldo, "4472C4" if saldo >= 0 else "FF6B6B"),
    ]

    row = 3
    for label, valor, cor in indicadores:
        ws_ind[f"A{row}"] = label
        ws_ind[f"A{row}"].font = bold
        ws_ind[f"A{row}"].fill = subheader_fill
        ws_ind[f"A{row}"].border = border

        ws_ind[f"B{row}"] = float(valor)
        ws_ind[f"B{row}"].number_format = currency_format
        ws_ind[f"B{row}"].font = Font(bold=True, color='FFFFFF')
        ws_ind[f"B{row}"].fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
        ws_ind[f"B{row}"].alignment = align_right
        ws_ind[f"B{row}"].border = border
        row += 1

    # === Gráficos ===
    ws_chart = wb.create_sheet('Graficos')
    ws_chart.column_dimensions['A'].width = 20
    ws_chart.column_dimensions['H'].width = 20

    # Linha: evolução do caixa (Resumo Mensal)
    if not resumo_mensal.empty:
        chart = LineChart()
        chart.title = 'Evolução do Caixa'
        chart.style = 13
        chart.y_axis.title = 'Valor (R$)'
        chart.x_axis.title = 'Período'

        data = Reference(ws_month, min_col=2, min_row=1, max_row=ws_month.max_row, max_col=4)
        cats = Reference(ws_month, min_col=1, min_row=2, max_row=ws_month.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 22
        ws_chart.add_chart(chart, 'A1')

    # Barras: entradas vs saídas (Resumo Anual)
    if not resumo_anual.empty:
        bchart = BarChart()
        bchart.title = 'Entradas vs Saídas (Anual)'
        bchart.y_axis.title = 'Valor (R$)'
        bchart.x_axis.title = 'Ano'
        vals = Reference(ws_year, min_col=2, min_row=1, max_row=ws_year.max_row)
        cats = Reference(ws_year, min_col=1, min_row=2, max_row=ws_year.max_row)
        bchart.add_data(vals, titles_from_data=True)
        bchart.set_categories(cats)
        bchart.height = 10
        bchart.width = 22
        ws_chart.add_chart(bchart, 'A20')

    # Pizza por categoria (top 10)
    if not df_raw.empty and 'categoria' in df_raw.columns and 'valor' in df_raw.columns:
        cat_counts = df_raw.groupby('categoria')['valor'].sum().reset_index()
        cat_counts = cat_counts.sort_values(by='valor', ascending=False).head(10)
        start_row = 40
        ws_chart[f"A{start_row}"] = "Categoria"
        ws_chart[f"B{start_row}"] = "Valor"
        ws_chart[f"A{start_row}"].font = bold
        ws_chart[f"B{start_row}"].font = bold
        ws_chart[f"A{start_row}"].fill = subheader_fill
        ws_chart[f"B{start_row}"].fill = subheader_fill

        for r_idx, row in enumerate(dataframe_to_rows(cat_counts, index=False, header=False), start=start_row + 1):
            ws_chart.cell(row=r_idx, column=1, value=row[0])
            cell_val = ws_chart.cell(row=r_idx, column=2, value=float(row[1]))
            cell_val.number_format = currency_format

        pchart = PieChart()
        labels = Reference(ws_chart, min_col=1, min_row=start_row + 1, max_row=start_row + len(cat_counts))
        data = Reference(ws_chart, min_col=2, min_row=start_row, max_row=start_row + len(cat_counts))
        pchart.add_data(data, titles_from_data=True)
        pchart.set_categories(labels)
        pchart.title = 'Distribuição por Categoria (Top 10)'
        pchart.height = 12
        pchart.width = 16
        ws_chart.add_chart(pchart, 'H1')

    wb.save(str(caminho_saida))
    return caminho_saida


def _build_styles():
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    zebra_fill = PatternFill(start_color='F7F9FB', end_color='F7F9FB', fill_type='solid')
    thin = Side(border_style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color='FFFFFF')
    bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    currency_format = 'R$ #,##0.00'
    date_format = 'yyyy-mm-dd'
    return {
        "header_fill": header_fill,
        "subheader_fill": subheader_fill,
        "zebra_fill": zebra_fill,
        "border": border,
        "header_font": header_font,
        "bold": bold,
        "align_center": align_center,
        "align_left": align_left,
        "align_right": align_right,
        "currency_format": currency_format,
        "date_format": date_format,
    }


def _coerce_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except Exception:
            try:
                return datetime.strptime(value, "%Y-%m-%d")
            except Exception:
                return value
    return value


def _aplicar_tabela(ws, headers, rows, column_widths, styles, currency_cols=None, date_cols=None, table_name="TabelaDados"):
    currency_cols = currency_cols or set()
    date_cols = date_cols or set()

    # Header
    for c_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["align_center"]
        cell.border = styles["border"]

    # Data
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            header = headers[c_idx - 1]
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = styles["border"]

            if header in date_cols:
                coerced = _coerce_date(value)
                cell.value = coerced
                cell.number_format = styles["date_format"]
                cell.alignment = styles["align_center"]
            elif header in currency_cols:
                try:
                    cell.value = float(value)
                except Exception:
                    pass
                cell.number_format = styles["currency_format"]
                cell.alignment = styles["align_right"]
            else:
                cell.alignment = styles["align_left"]

        if r_idx % 2 == 0:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = styles["zebra_fill"]

    # Widths
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    try:
        tabela = Table(displayName=table_name, ref=ws.auto_filter.ref)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tabela)
    except Exception:
        pass


def gerar_planilha_clientes(clientes: list, caminho_saida: Path):
    styles = _build_styles()
    wb = Workbook()

    headers = ["ID", "Tipo", "Nome", "Documento", "Email", "Telefone", "Cidade", "Status", "Data Cadastro"]
    rows = []
    for c in clientes:
        tipo = c.get("tipo_pessoa", "")
        tipo_label = "Fisica" if str(tipo).lower() == "fisica" else "Juridica"
        rows.append([
            c.get("id", ""),
            tipo_label,
            c.get("nome", ""),
            c.get("documento", ""),
            c.get("email", ""),
            c.get("telefone", ""),
            c.get("cidade", ""),
            c.get("status", ""),
            c.get("data_cadastro", ""),
        ])

    ws = wb.active
    ws.title = "Clientes"
    _aplicar_tabela(
        ws,
        headers,
        rows,
        [8, 10, 26, 18, 28, 16, 16, 10, 20],
        styles,
        currency_cols=set(),
        date_cols={"Data Cadastro"},
        table_name="TabelaClientes",
    )

    # Resumo
    ws_sum = wb.create_sheet("Resumo")
    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["D"].width = 18
    ws_sum.column_dimensions["E"].width = 12

    ws_sum["A1"] = "RELATORIO EXECUTIVO - CLIENTES"
    ws_sum["A1"].font = styles["header_font"]
    ws_sum["A1"].fill = styles["header_fill"]
    ws_sum["A1"].alignment = styles["align_center"]
    ws_sum.merge_cells("A1:E1")

    total = len(rows)
    status_counts = {"ativo": 0, "inativo": 0, "suspenso": 0}
    tipo_counts = {"Fisica": 0, "Juridica": 0}
    for r in rows:
        status = str(r[7]).lower()
        if status in status_counts:
            status_counts[status] += 1
        tipo_counts[r[1]] = tipo_counts.get(r[1], 0) + 1

    ws_sum["A3"] = "Total de Clientes"
    ws_sum["B3"] = total
    ws_sum["A3"].font = styles["bold"]
    ws_sum["B3"].font = styles["bold"]

    ws_sum["A5"] = "Status"
    ws_sum["A5"].font = styles["bold"]
    ws_sum["A5"].fill = styles["subheader_fill"]
    ws_sum["B5"].fill = styles["subheader_fill"]
    ws_sum["A6"] = "Ativo"
    ws_sum["B6"] = status_counts["ativo"]
    ws_sum["A7"] = "Inativo"
    ws_sum["B7"] = status_counts["inativo"]
    ws_sum["A8"] = "Suspenso"
    ws_sum["B8"] = status_counts["suspenso"]

    ws_sum["D5"] = "Tipo Pessoa"
    ws_sum["D5"].font = styles["bold"]
    ws_sum["D5"].fill = styles["subheader_fill"]
    ws_sum["E5"].fill = styles["subheader_fill"]
    ws_sum["D6"] = "Fisica"
    ws_sum["E6"] = tipo_counts.get("Fisica", 0)
    ws_sum["D7"] = "Juridica"
    ws_sum["E7"] = tipo_counts.get("Juridica", 0)

    # Charts
    pie = PieChart()
    pie.title = "Status"
    data = Reference(ws_sum, min_col=2, min_row=6, max_row=8)
    labels = Reference(ws_sum, min_col=1, min_row=6, max_row=8)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 8
    pie.width = 12
    ws_sum.add_chart(pie, "G2")

    bchart = BarChart()
    bchart.title = "Tipo Pessoa"
    data = Reference(ws_sum, min_col=5, min_row=6, max_row=7)
    labels = Reference(ws_sum, min_col=4, min_row=6, max_row=7)
    bchart.add_data(data, titles_from_data=False)
    bchart.set_categories(labels)
    bchart.height = 8
    bchart.width = 12
    ws_sum.add_chart(bchart, "G15")

    wb.save(str(caminho_saida))
    return caminho_saida


def gerar_planilha_fornecedores(fornecedores: list, caminho_saida: Path):
    styles = _build_styles()
    wb = Workbook()

    headers = ["ID", "Tipo", "Nome", "CPF/CNPJ", "Telefone", "Email", "Cidade", "Status"]
    rows = []
    for f in fornecedores:
        tipo_val = getattr(f, "tipo", "")
        tipo = tipo_val.value if hasattr(tipo_val, "value") else str(tipo_val)
        tipo_label = "Fisica" if tipo.lower() == "fisica" else "Juridica"
        rows.append([
            getattr(f, "id", ""),
            tipo_label,
            getattr(f, "nome", ""),
            getattr(f, "cpf_cnpj", ""),
            getattr(f, "telefone", ""),
            getattr(f, "email", ""),
            getattr(f, "cidade", ""),
            getattr(f, "status", ""),
        ])

    ws = wb.active
    ws.title = "Fornecedores"
    _aplicar_tabela(
        ws,
        headers,
        rows,
        [8, 10, 28, 18, 16, 28, 16, 10],
        styles,
        currency_cols=set(),
        date_cols=set(),
        table_name="TabelaFornecedores",
    )

    ws_sum = wb.create_sheet("Resumo")
    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["D"].width = 18
    ws_sum.column_dimensions["E"].width = 12

    ws_sum["A1"] = "RELATORIO EXECUTIVO - FORNECEDORES"
    ws_sum["A1"].font = styles["header_font"]
    ws_sum["A1"].fill = styles["header_fill"]
    ws_sum["A1"].alignment = styles["align_center"]
    ws_sum.merge_cells("A1:E1")

    total = len(rows)
    status_counts = {"ativo": 0, "inativo": 0}
    tipo_counts = {"Fisica": 0, "Juridica": 0}
    for r in rows:
        status = str(r[7]).lower()
        if status in status_counts:
            status_counts[status] += 1
        tipo_counts[r[1]] = tipo_counts.get(r[1], 0) + 1

    ws_sum["A3"] = "Total de Fornecedores"
    ws_sum["B3"] = total
    ws_sum["A3"].font = styles["bold"]
    ws_sum["B3"].font = styles["bold"]

    ws_sum["A5"] = "Status"
    ws_sum["A5"].font = styles["bold"]
    ws_sum["A5"].fill = styles["subheader_fill"]
    ws_sum["B5"].fill = styles["subheader_fill"]
    ws_sum["A6"] = "Ativo"
    ws_sum["B6"] = status_counts["ativo"]
    ws_sum["A7"] = "Inativo"
    ws_sum["B7"] = status_counts["inativo"]

    ws_sum["D5"] = "Tipo"
    ws_sum["D5"].font = styles["bold"]
    ws_sum["D5"].fill = styles["subheader_fill"]
    ws_sum["E5"].fill = styles["subheader_fill"]
    ws_sum["D6"] = "Fisica"
    ws_sum["E6"] = tipo_counts.get("Fisica", 0)
    ws_sum["D7"] = "Juridica"
    ws_sum["E7"] = tipo_counts.get("Juridica", 0)

    pie = PieChart()
    pie.title = "Status"
    data = Reference(ws_sum, min_col=2, min_row=6, max_row=7)
    labels = Reference(ws_sum, min_col=1, min_row=6, max_row=7)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 8
    pie.width = 12
    ws_sum.add_chart(pie, "G2")

    bchart = BarChart()
    bchart.title = "Tipo"
    data = Reference(ws_sum, min_col=5, min_row=6, max_row=7)
    labels = Reference(ws_sum, min_col=4, min_row=6, max_row=7)
    bchart.add_data(data, titles_from_data=False)
    bchart.set_categories(labels)
    bchart.height = 8
    bchart.width = 12
    ws_sum.add_chart(bchart, "G15")

    wb.save(str(caminho_saida))
    return caminho_saida


def gerar_planilha_funcionarios(funcionarios: list, caminho_saida: Path):
    styles = _build_styles()
    wb = Workbook()

    headers = ["ID", "Nome", "CPF", "Cargo", "Email", "Telefone", "Salario", "Admissao", "Status"]
    rows = []
    for f in funcionarios:
        rows.append([
            f.get("id", ""),
            f.get("nome", ""),
            f.get("cpf", ""),
            f.get("cargo", ""),
            f.get("email", ""),
            f.get("telefone", ""),
            f.get("salario", ""),
            f.get("data_admissao", ""),
            f.get("status", ""),
        ])

    ws = wb.active
    ws.title = "Funcionarios"
    _aplicar_tabela(
        ws,
        headers,
        rows,
        [8, 26, 16, 20, 28, 16, 12, 12, 12],
        styles,
        currency_cols={"Salario"},
        date_cols={"Admissao"},
        table_name="TabelaFuncionarios",
    )

    ws_sum = wb.create_sheet("Resumo")
    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["D"].width = 22
    ws_sum.column_dimensions["E"].width = 12

    ws_sum["A1"] = "RELATORIO EXECUTIVO - FUNCIONARIOS"
    ws_sum["A1"].font = styles["header_font"]
    ws_sum["A1"].fill = styles["header_fill"]
    ws_sum["A1"].alignment = styles["align_center"]
    ws_sum.merge_cells("A1:E1")

    total = len(rows)
    status_counts = {"ativo": 0, "inativo": 0, "licenca": 0, "desligado": 0}
    cargo_counts = {}
    for r in rows:
        status = str(r[8]).lower()
        if status in status_counts:
            status_counts[status] += 1
        cargo = str(r[3]).strip() or "Nao informado"
        cargo_counts[cargo] = cargo_counts.get(cargo, 0) + 1

    ws_sum["A3"] = "Total de Funcionarios"
    ws_sum["B3"] = total
    ws_sum["A3"].font = styles["bold"]
    ws_sum["B3"].font = styles["bold"]

    ws_sum["A5"] = "Status"
    ws_sum["A5"].font = styles["bold"]
    ws_sum["A5"].fill = styles["subheader_fill"]
    ws_sum["B5"].fill = styles["subheader_fill"]
    ws_sum["A6"] = "Ativo"
    ws_sum["B6"] = status_counts["ativo"]
    ws_sum["A7"] = "Inativo"
    ws_sum["B7"] = status_counts["inativo"]
    ws_sum["A8"] = "Licenca"
    ws_sum["B8"] = status_counts["licenca"]
    ws_sum["A9"] = "Desligado"
    ws_sum["B9"] = status_counts["desligado"]

    # Top cargos
    ws_sum["D5"] = "Top Cargos"
    ws_sum["D5"].font = styles["bold"]
    ws_sum["D5"].fill = styles["subheader_fill"]
    ws_sum["E5"].fill = styles["subheader_fill"]
    top_cargos = sorted(cargo_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    row = 6
    for cargo, qtd in top_cargos:
        ws_sum[f"D{row}"] = cargo
        ws_sum[f"E{row}"] = qtd
        row += 1

    pie = PieChart()
    pie.title = "Status"
    data = Reference(ws_sum, min_col=2, min_row=6, max_row=9)
    labels = Reference(ws_sum, min_col=1, min_row=6, max_row=9)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 8
    pie.width = 12
    ws_sum.add_chart(pie, "G2")

    if top_cargos:
        bchart = BarChart()
        bchart.title = "Cargos"
        data = Reference(ws_sum, min_col=5, min_row=6, max_row=5 + len(top_cargos))
        labels = Reference(ws_sum, min_col=4, min_row=6, max_row=5 + len(top_cargos))
        bchart.add_data(data, titles_from_data=False)
        bchart.set_categories(labels)
        bchart.height = 10
        bchart.width = 16
        ws_sum.add_chart(bchart, "G15")

    wb.save(str(caminho_saida))
    return caminho_saida
