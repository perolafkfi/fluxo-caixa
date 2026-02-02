"""
Tela de Relatórios e Impressão
Interface para visualizar e imprimir relatórios dentro do sistema
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from datetime import datetime, timedelta
from pathlib import Path
import threading
import matplotlib.pyplot as plt
import matplotlib
from matplotlib.ticker import FuncFormatter
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

from app.services.relatorios import GeradorRelatorios

# Configurar matplotlib para tema claro
matplotlib.rcParams['figure.facecolor'] = '#ffffff'
matplotlib.rcParams['axes.facecolor'] = '#f5f5f5'
matplotlib.rcParams['text.color'] = '#333333'
matplotlib.rcParams['axes.labelcolor'] = '#333333'
matplotlib.rcParams['xtick.color'] = '#333333'
matplotlib.rcParams['ytick.color'] = '#333333'
matplotlib.rcParams['font.family'] = 'sans-serif'
matplotlib.rcParams['font.size'] = 10


class TelaRelatorios:
    """Interface de relatórios e impressão com gráficos"""
    
    # Paleta de cores profissional
    CORES_CATEGORIAS = {
        'Receita': '#27ae60',      # Verde
        'Variável': '#e74c3c',     # Vermelho
        'Fixa': '#f39c12',         # Laranja
        'Pessoal': '#8e44ad',      # Roxo
        'Receita': '#27ae60',
        'Despesa': '#e74c3c',
        'entrada': '#27ae60',
        'saida': '#e74c3c'
    }
    
    CORES_GRADIENTE = ['#3498db', '#27ae60', '#e74c3c', '#f39c12', '#8e44ad', '#1abc9c']
    
    def __init__(self, parent, gerador_relatorios: GeradorRelatorios):
        self.parent = parent
        self.gerador = gerador_relatorios
        self.relatorio_atual = ""
        self._style = ttk.Style()
    
    def criar_interface(self, frame_principal):
        """Cria interface da tela com gráficos"""
        frame_principal.grid_rowconfigure(0, weight=0)
        frame_principal.grid_rowconfigure(1, weight=0, minsize=90)
        frame_principal.grid_rowconfigure(2, weight=1)
        frame_principal.grid_columnconfigure(0, weight=1)
        
        # === LINHA 0: Filtros ===
        self._criar_painel_filtros(frame_principal)
        
        # === LINHA 1: Resumo Executivo ===
        self._criar_resumo_executivo(frame_principal)
        
        # === LINHA 2: Gráficos em Grid (2x2) ===
        self._criar_painel_graficos(frame_principal)
        
        # Carregar dados inicialmente
        self.atualizar_relatorios()
    
    def _criar_painel_filtros(self, parent):
        """Cria painel de filtros compacto"""
        frame_filtros = ttk.LabelFrame(parent, text="Período", padding=8)
        frame_filtros.grid(row=0, column=0, sticky='ew', padx=10, pady=5)

        ttk.Label(frame_filtros, text="De:").pack(side=tk.LEFT, padx=5)
        self.entry_data_inicio = ttk.Entry(frame_filtros, width=12)
        self.entry_data_inicio.pack(side=tk.LEFT, padx=5)
        self.entry_data_inicio.insert(0, (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))

        ttk.Label(frame_filtros, text="Até:").pack(side=tk.LEFT, padx=5)
        self.entry_data_fim = ttk.Entry(frame_filtros, width=12)
        self.entry_data_fim.pack(side=tk.LEFT, padx=5)
        self.entry_data_fim.insert(0, datetime.now().strftime("%Y-%m-%d"))

        ttk.Button(frame_filtros, text="↻ Atualizar", 
                  command=self.atualizar_relatorios).pack(side=tk.LEFT, padx=10)
        ttk.Button(frame_filtros, text="✕ Limpar", 
                  command=self.limpar_filtros).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_filtros, text="🖨️ Imprimir", 
              command=self.imprimir_relatorio).pack(side=tk.LEFT, padx=6)
        ttk.Button(frame_filtros, text="📊 Excel", 
              command=self.exportar_relatorio_profissional).pack(side=tk.LEFT, padx=6)

    def _criar_resumo_executivo(self, parent):
        """Cria cards do resumo executivo"""
        frame_resumo = ttk.Frame(parent)
        frame_resumo.grid(row=1, column=0, sticky='ew', padx=10, pady=10)
        frame_resumo.grid_columnconfigure((0, 1, 2), weight=1)

        # Card Receitas
        self._criar_card(frame_resumo, 0, "RECEITAS", "label_receitas", "#27ae60")
        # Card Despesas
        self._criar_card(frame_resumo, 1, "DESPESAS", "label_despesas", "#e74c3c")
        # Card Saldo
        self._criar_card(frame_resumo, 2, "SALDO", "label_saldo", "#3498db")

    def _criar_card(self, parent, col, titulo, label_var, cor):
        """Cria um card de resumo"""
        frame_card = ttk.Frame(parent, relief=tk.RIDGE, borderwidth=1, padding=8)
        frame_card.grid(row=0, column=col, sticky='nsew', padx=5, pady=5)

        title_style = f"{label_var}.Title.TLabel"
        value_style = f"{label_var}.Value.TLabel"
        self._style.configure(title_style, font=("Segoe UI", 10, "bold"))
        self._style.configure(value_style, font=("Segoe UI", 16, "bold"), foreground=cor)

        ttk.Label(frame_card, text=titulo, style=title_style).pack(pady=(6, 3))
        
        label = ttk.Label(frame_card, text="R$ 0,00", style=value_style)
        label.pack(pady=(0, 6))
        
        # Armazenar referência do label
        setattr(self, label_var, label)

    def _criar_painel_graficos(self, parent):
        """Cria painel com gráficos em grid 2x2"""
        frame_graficos = ttk.Frame(parent)
        frame_graficos.grid(row=2, column=0, sticky='nsew', padx=10, pady=10)
        parent.grid_rowconfigure(2, weight=1)

        frame_graficos.grid_rowconfigure((0, 1), weight=1)
        frame_graficos.grid_columnconfigure((0, 1), weight=1)

        # Gráfico 1: Categoria (Superior Esquerdo)
        self.frame_grafico_categoria = ttk.LabelFrame(frame_graficos, text="Total por Categoria", padding=5)
        self.frame_grafico_categoria.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)

        # Gráfico 2: Subcategoria (Superior Direito)
        self.frame_grafico_subcategoria = ttk.LabelFrame(frame_graficos, text="Total por Tipo", padding=5)
        self.frame_grafico_subcategoria.grid(row=0, column=1, sticky='nsew', padx=5, pady=5)

        # Gráfico 3: Receita vs Despesa (Inferior Esquerdo)
        self.frame_grafico_despesa = ttk.LabelFrame(frame_graficos, text="Receita vs Despesa", padding=5)
        self.frame_grafico_despesa.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)

        # Tabela 4: Detalhes (Inferior Direito)
        frame_tabela = ttk.LabelFrame(frame_graficos, text="Detalhamento", padding=5)
        frame_tabela.grid(row=1, column=1, sticky='nsew', padx=5, pady=5)

        self._criar_tabela_detalhes(frame_tabela)

    def _criar_tabela_detalhes(self, parent):
        """Cria tabela de detalhamento agregado"""
        colunas = ("Tipo", "Categoria", "Valor")
        self.tree_detalhes = ttk.Treeview(parent, columns=colunas, height=12, show="tree headings")

        # Configurar colunas
        self.tree_detalhes.column("#0", width=0, stretch=tk.NO)
        self.tree_detalhes.column("Tipo", anchor=tk.CENTER, width=80)
        self.tree_detalhes.column("Categoria", anchor=tk.W, width=120)
        self.tree_detalhes.column("Valor", anchor=tk.E, width=70)

        # Headers
        for col in colunas:
            self.tree_detalhes.heading(col, text=col)

        # Scrollbar
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self.tree_detalhes.yview)
        self.tree_detalhes.configure(yscroll=scrollbar.set)

        self.tree_detalhes.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def _limpar_canvas(self, frame):
        """Remove todos os widgets do frame"""
        for widget in frame.winfo_children():
            widget.destroy()

    def atualizar_relatorios(self):
        """Atualiza todos os relatórios com um único ponto de entrada"""
        try:
            filtros = self.obter_filtros()
            
            # Gerar gráficos
            self.plotar_grafico_categoria(filtros)
            self.plotar_grafico_tipo(filtros)
            self.plotar_grafico_receita_despesa(filtros)
            self.carregar_tabela_detalhes(filtros)
            
            # Atualizar resumo
            self._atualizar_resumo(filtros)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao atualizar relatórios: {str(e)}")

    def obter_filtros(self) -> dict:
        """Obtém filtros de período"""
        filtros = {}
        if self.entry_data_inicio.get():
            filtros['data_inicio'] = self.entry_data_inicio.get()
        if self.entry_data_fim.get():
            filtros['data_fim'] = self.entry_data_fim.get()
        return filtros

    def _atualizar_resumo(self, filtros: dict):
        """Atualiza labels do resumo executivo com formatação"""
        try:
            resumo = self.gerador.gerar_resumo(filtros)
            total_receitas = float(resumo.get('total_receitas', 0))
            total_despesas = float(resumo.get('total_despesas', 0))
            saldo = float(resumo.get('saldo', 0))

            self.label_receitas.config(text=self._formatar_moeda(total_receitas))
            self.label_despesas.config(text=self._formatar_moeda(total_despesas))

            cor = "#27ae60" if saldo >= 0 else "#e74c3c"
            self._style.configure("label_saldo.Value.TLabel", foreground=cor)
            self.label_saldo.config(text=self._formatar_moeda(saldo))
        except Exception as e:
            print(f"Erro ao atualizar resumo: {e}")

    def _formatar_moeda(self, valor: float) -> str:
        try:
            texto = f"{valor:,.2f}"
        except Exception:
            texto = "0.00"
        texto = texto.replace(",", "X").replace(".", ",").replace("X", ".")
        return f"R$ {texto}"

    def _formatar_moeda_curta(self, valor: float) -> str:
        abs_valor = abs(valor)
        if abs_valor >= 1_000_000:
            texto = f"{abs_valor/1_000_000:.1f}M"
        elif abs_valor >= 1_000:
            texto = f"{abs_valor/1_000:.1f}k"
        else:
            texto = f"{abs_valor:.0f}"
        texto = texto.replace(".", ",")
        return texto

    def _normalizar_tipo(self, tipo_valor):
        tipo_normalizado = str(tipo_valor).strip().lower()
        if tipo_normalizado in ("receita", "entrada"):
            return "Receita"
        if tipo_normalizado in ("despesa", "saida", "saída"):
            return "Despesa"
        return None

    def plotar_grafico_categoria(self, filtros: dict):
        """Gráfico de barras por categoria com tema claro"""
        try:
            lancamentos = self.gerador.obter_lancamentos_por_periodo(
                filtros.get('data_inicio', ''),
                filtros.get('data_fim', '')
            )
            
            if not lancamentos:
                self._limpar_canvas(self.frame_grafico_categoria)
                return

            # Agrupar por categoria
            dados = {}
            for lanc in lancamentos:
                if isinstance(lanc, dict):
                    cat = lanc.get('categoria') or "Sem categoria"
                    valor = lanc.get('valor', 0)
                else:
                    cat_obj = getattr(lanc, 'categoria', None)
                    cat = cat_obj.value if cat_obj else "Sem categoria"
                    valor = getattr(lanc, 'valor', 0)
                dados[cat] = dados.get(cat, 0) + abs(valor)

            if not dados:
                self._limpar_canvas(self.frame_grafico_categoria)
                return

            itens = sorted(dados.items(), key=lambda x: x[1], reverse=True)
            categorias = [i[0] for i in itens]
            valores = [i[1] for i in itens]

            self._limpar_canvas(self.frame_grafico_categoria)

            fig = Figure(figsize=(10, 5.5), dpi=100)
            ax = fig.add_subplot(111)

            cores = [self.CORES_GRADIENTE[i % len(self.CORES_GRADIENTE)] for i in range(len(categorias))]
            ax.bar(
                range(len(categorias)),
                valores,
                color=cores,
                edgecolor='#ffffff',
                linewidth=1.2,
                alpha=0.9,
                width=0.6
            )

            # Adicionar margem no topo para os números não serem cortados
            max_valor = max(valores) if valores else 0
            ax.set_ylim(0, max_valor * 1.15 if max_valor else 1)
            
            ax.set_ylabel('Valor (R$)', fontsize=10, fontweight='bold')
            ax.set_xticks(range(len(categorias)))
            ax.set_xticklabels([c.replace('_', ' ').title()[:12] for c in categorias], rotation=20, ha='right', fontsize=9)
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: self._formatar_moeda_curta(x)))
            ax.grid(axis='y', alpha=0.25, linestyle='--', linewidth=1)
            ax.set_axisbelow(True)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

            fig.subplots_adjust(bottom=0.25, left=0.1, right=0.95, top=0.9)
            canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico_categoria)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        except Exception as e:
            print(f"Erro ao plotar gráfico de categoria: {e}")

    def plotar_grafico_tipo(self, filtros: dict):
        """Gráfico horizontal por tipo (receita/despesa)"""
        try:
            lancamentos = self.gerador.obter_lancamentos_por_periodo(
                filtros.get('data_inicio', ''),
                filtros.get('data_fim', '')
            )
            
            if not lancamentos:
                self._limpar_canvas(self.frame_grafico_subcategoria)
                return

            # Agrupar por tipo
            dados = {}
            for lanc in lancamentos:
                tipo_valor = lanc.get('tipo') if isinstance(lanc, dict) else lanc.tipo.value
                valor = lanc.get('valor') if isinstance(lanc, dict) else lanc.valor
                tipo = self._normalizar_tipo(tipo_valor)
                if not tipo:
                    continue
                dados[tipo] = dados.get(tipo, 0) + abs(valor)

            if not dados:
                self._limpar_canvas(self.frame_grafico_subcategoria)
                return

            itens = sorted(dados.items(), key=lambda x: x[1], reverse=True)
            tipos = [i[0] for i in itens]
            valores = [i[1] for i in itens]

            self._limpar_canvas(self.frame_grafico_subcategoria)

            fig = Figure(figsize=(10, 5.5), dpi=100)
            ax = fig.add_subplot(111)

            cores = [self.CORES_CATEGORIAS.get(t, '#3498db') for t in tipos]
            ax.barh(
                tipos,
                valores,
                color=cores,
                edgecolor='#ffffff',
                linewidth=1.2,
                alpha=0.9,
                height=0.6
            )

            # Adicionar margem na direita para os números não serem cortados
            max_valor = max(valores) if valores else 0
            ax.set_xlim(0, max_valor * 1.2 if max_valor else 1)
            
            ax.set_xlabel('Valor (R$)', fontsize=10, fontweight='bold')
            ax.xaxis.set_major_formatter(FuncFormatter(lambda x, pos: self._formatar_moeda_curta(x)))
            ax.grid(axis='x', alpha=0.25, linestyle='--', linewidth=1)
            ax.set_axisbelow(True)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

            fig.subplots_adjust(bottom=0.15, left=0.15, right=0.85, top=0.9)
            canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico_subcategoria)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        except Exception as e:
            print(f"Erro ao plotar gráfico de tipo: {e}")

    def plotar_grafico_receita_despesa(self, filtros: dict):
        """Gráfico de pizza para distribuição receita vs despesa"""
        try:
            lancamentos = self.gerador.obter_lancamentos_por_periodo(
                filtros.get('data_inicio', ''),
                filtros.get('data_fim', '')
            )
            
            if not lancamentos:
                self._limpar_canvas(self.frame_grafico_despesa)
                return

            # Agrupar receitas e despesas
            totais = {"Receita": 0.0, "Despesa": 0.0}
            for lanc in lancamentos:
                tipo_valor = lanc.get('tipo') if isinstance(lanc, dict) else lanc.tipo.value
                valor = lanc.get('valor') if isinstance(lanc, dict) else lanc.valor
                tipo = self._normalizar_tipo(tipo_valor)
                if not tipo:
                    continue
                totais[tipo] += abs(valor or 0)

            totais = {k: v for k, v in totais.items() if v > 0}
            if not totais:
                self._limpar_canvas(self.frame_grafico_despesa)
                return
            tipos = list(totais.keys())
            valores = list(totais.values())
            cores = [self.CORES_CATEGORIAS.get(t, '#3498db') for t in tipos]

            self._limpar_canvas(self.frame_grafico_despesa)

            fig = Figure(figsize=(10, 6), dpi=100)
            ax = fig.add_subplot(111)

            # Criar gráfico de pizza em formato donut (visual limpo)
            wedges, _ = ax.pie(
                valores,
                labels=None,
                autopct=None,
                colors=cores,
                startangle=90,
                explode=[0.05 if v == max(valores) else 0.02 for v in valores],
                wedgeprops={'edgecolor': '#ffffff', 'linewidth': 2.0, 'width': 0.45}
            )

            total_geral = sum(valores) if valores else 0
            ax.text(0, 0.05, self._formatar_moeda(total_geral), ha='center', va='center', fontsize=12, fontweight='bold')
            ax.text(0, -0.12, 'Total', ha='center', va='center', fontsize=10, color='#555555')
            
            # Adicionar legenda com valores e percentuais
            legend_labels = []
            for i in range(len(tipos)):
                perc = (valores[i] / total_geral * 100) if total_geral else 0
                legend_labels.append(f"{tipos[i]}: {self._formatar_moeda(valores[i])} ({perc:.1f}%)")
            ax.legend(legend_labels, loc='upper left', bbox_to_anchor=(0.72, 1), fontsize=10,
                     title='Detalhes', title_fontsize=12, framealpha=0.95)
            
            fig.subplots_adjust(bottom=0.1, left=0.05, right=0.72, top=0.9)
            canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico_despesa)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        except Exception as e:
            print(f"Erro ao plotar gráfico de receita/despesa: {e}")

    def carregar_tabela_detalhes(self, filtros: dict):
        """Carrega tabela de detalhes com formatação"""
        try:
            # Limpar tabela anterior
            for item in self.tree_detalhes.get_children():
                self.tree_detalhes.delete(item)

            lancamentos = self.gerador.obter_lancamentos_por_periodo(
                filtros.get('data_inicio', ''),
                filtros.get('data_fim', '')
            )
            
            resumo_detalhe = {}

            # Agregar dados
            for lanc in lancamentos:
                tipo = lanc.get('tipo') if isinstance(lanc, dict) else lanc.tipo.value
                categoria = lanc.get('categoria') if isinstance(lanc, dict) else lanc.categoria.value
                valor = lanc.get('valor') if isinstance(lanc, dict) else lanc.valor
                chave = (tipo, categoria)
                if chave not in resumo_detalhe:
                    resumo_detalhe[chave] = 0
                resumo_detalhe[chave] += valor

            # Inserir na tabela
            for (tipo, categoria), total in sorted(resumo_detalhe.items()):
                self.tree_detalhes.insert("", tk.END, values=(
                    tipo.title(),
                    categoria.replace('_', ' ').title(),
                    self._formatar_moeda(abs(total))
                ))
        except Exception as e:
            print(f"Erro ao carregar tabela: {e}")

    def limpar_filtros(self):
        """Limpa filtros"""
        self.entry_data_inicio.delete(0, tk.END)
        self.entry_data_fim.delete(0, tk.END)
        self.entry_data_inicio.insert(0, (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
        self.entry_data_fim.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.atualizar_relatorios()

    def exportar_relatorio(self):
        """Exporta relatório de resumo para Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        try:
            arquivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
                initialfile=f"Relatorio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not arquivo:
                return
            
            # Obter datas de filtro
            try:
                data_inicio = datetime.strptime(self.entry_data_inicio.get(), "%Y-%m-%d")
                data_fim = datetime.strptime(self.entry_data_fim.get(), "%Y-%m-%d")
            except ValueError:
                data_inicio = datetime.now() - timedelta(days=30)
                data_fim = datetime.now()
            
            # Obter dados dos gráficos
            lancamentos = self.gerador.obter_lancamentos_filtrados({})
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório"
            
            # === SEÇÃO 1: RESUMO ===
            ws['A1'] = "RELATÓRIO DE FLUXO DE CAIXA"
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            ws.merge_cells('A1:D1')
            
            ws['A2'] = f"Período: {self.entry_data_inicio.get()} a {self.entry_data_fim.get()}"
            ws['A2'].font = Font(size=10)
            ws.merge_cells('A2:D2')
            
            row = 4
            
            # Totais
            total_receitas = self.gerador.calcular_total_receitas({})
            total_despesas = self.gerador.calcular_total_despesas({})
            saldo = total_receitas - total_despesas
            
            ws[f'A{row}'] = "RECEITAS"
            ws[f'B{row}'] = f"R$ {total_receitas:,.2f}"
            ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
            row += 1
            
            ws[f'A{row}'] = "DESPESAS"
            ws[f'B{row}'] = f"R$ {total_despesas:,.2f}"
            ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
            row += 1
            
            ws[f'A{row}'] = "SALDO"
            ws[f'B{row}'] = f"R$ {saldo:,.2f}"
            ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
            cor_saldo = "1abc9c" if saldo >= 0 else "e74c3c"
            ws[f'A{row}'].fill = PatternFill(start_color=cor_saldo, end_color=cor_saldo, fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="e6f7f5", end_color="e6f7f5", fill_type="solid")
            row += 3
            
            # === SEÇÃO 2: TOTAIS POR CATEGORIA ===
            ws[f'A{row}'] = "TOTAIS POR CATEGORIA"
            ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            row += 1
            
            categorias_totais = self.gerador.totais_por_categoria({})
            
            ws[f'A{row}'] = "Categoria"
            ws[f'B{row}'] = "Valor"
            for col in ['A', 'B']:
                ws[f'{col}{row}'].font = Font(bold=True)
                ws[f'{col}{row}'].fill = PatternFill(start_color="bdc3c7", end_color="bdc3c7", fill_type="solid")
            row += 1
            
            for categoria, valor in sorted(categorias_totais.items()):
                ws[f'A{row}'] = str(categoria).replace('_', ' ').title()
                ws[f'B{row}'] = f"R$ {abs(valor):,.2f}"
                ws[f'B{row}'].alignment = Alignment(horizontal="right")
                row += 1
            
            # === SEÇÃO 3: DETALHAMENTO ===
            row += 2
            ws[f'A{row}'] = "DETALHAMENTO COMPLETO"
            ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            row += 1
            
            headers_detalhe = ["Data", "Tipo", "Categoria", "Descrição", "Valor"]
            for col_idx, header in enumerate(headers_detalhe, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            row += 1
            
            for lanc in lancamentos:
                tipo = lanc.get('tipo') if isinstance(lanc, dict) else lanc.tipo.value
                data = lanc.get('data') if isinstance(lanc, dict) else lanc.data
                categoria = lanc.get('categoria') if isinstance(lanc, dict) else lanc.categoria.value
                descricao = lanc.get('descricao') if isinstance(lanc, dict) else lanc.descricao
                valor = lanc.get('valor') if isinstance(lanc, dict) else lanc.valor
                
                ws.cell(row=row, column=1).value = data
                ws.cell(row=row, column=2).value = tipo.title()
                ws.cell(row=row, column=3).value = categoria.replace('_', ' ').title()
                ws.cell(row=row, column=4).value = descricao
                ws.cell(row=row, column=5).value = f"R$ {abs(valor):,.2f}"
                ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
                row += 1
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 15
            
            wb.save(arquivo)
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso!\n\n{arquivo}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")

    def imprimir_relatorio(self):
        """Gera PDF profissional via serviço e imprime (ou abre)."""
        try:
            from app.services.impressao import gerar_pdf_relatorio, imprimir_pdf_windows

            filtros = self.obter_filtros()
            periodo = f"{filtros.get('data_inicio','')} a {filtros.get('data_fim','')}"
            lancamentos = self.gerador.obter_lancamentos_por_periodo(filtros.get('data_inicio',''), filtros.get('data_fim',''))
            resumo = self.gerador.gerar_resumo(filtros)

            totais = {
                'entradas': resumo.get('total_receitas', 0),
                'saidas': resumo.get('total_despesas', 0),
                'saldo': resumo.get('saldo', 0)
            }

            caminho = gerar_pdf_relatorio('Fluxo de Caixa Profissional', periodo, lancamentos, totais)
            imprimir_pdf_windows(caminho)
            messagebox.showinfo('Impressão', f'Relatório gerado: {caminho}')
        except Exception as e:
            messagebox.showerror('Erro', f'Erro ao imprimir relatório: {e}')

    def exportar_relatorio_profissional(self):
        """Gera exportação Excel profissional por serviço."""
        try:
            from app.services.export_excel_profissional import gerar_planilha_profissional
            from tkinter import filedialog
            import pandas as pd

            arquivo = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files','*.xlsx')],
                                                   initialfile=f"Relatorio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            if not arquivo:
                return

            filtros = self.obter_filtros()
            lancamentos = self.gerador.obter_lancamentos_por_periodo(filtros.get('data_inicio',''), filtros.get('data_fim',''))

            df = pd.DataFrame(lancamentos)
            # Resumo mensal
            if not df.empty:
                df['data'] = pd.to_datetime(df['data'])
                df['ano_mes'] = df['data'].dt.to_period('M').astype(str)
                entradas = df[df['tipo'] == 'Receita'].groupby('ano_mes')['valor'].sum().rename('entradas')
                saidas = df[df['tipo'] == 'Despesa'].groupby('ano_mes')['valor'].sum().rename('saidas')
                resumo_mensal = pd.concat([entradas, saidas], axis=1).fillna(0).reset_index()
                resumo_mensal['saldo'] = resumo_mensal['entradas'] - resumo_mensal['saidas']
            else:
                resumo_mensal = pd.DataFrame(columns=['ano_mes','entradas','saidas','saldo'])

            # Resumo anual
            if not df.empty:
                df['ano'] = df['data'].dt.year
                entradas_a = df[df['tipo'] == 'Receita'].groupby('ano')['valor'].sum().rename('entradas')
                saidas_a = df[df['tipo'] == 'Despesa'].groupby('ano')['valor'].sum().rename('saidas')
                resumo_anual = pd.concat([entradas_a, saidas_a], axis=1).fillna(0).reset_index()
                resumo_anual['saldo'] = resumo_anual['entradas'] - resumo_anual['saidas']
            else:
                resumo_anual = pd.DataFrame(columns=['ano','entradas','saidas','saldo'])

            gerar_planilha_profissional(lancamentos, resumo_mensal, resumo_anual, Path(arquivo))
            messagebox.showinfo('Sucesso', f'Exportação profissional salva em:\n{arquivo}')
        except Exception as e:
            messagebox.showerror('Erro', f'Erro ao exportar Excel profissional: {e}')

